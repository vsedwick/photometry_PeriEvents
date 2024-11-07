import pandas as pd
import os
import openpyxl

compile_directory = r"E:\Photometry-Fall2022\Final Analysis\REANALYSIS_nolimit\Re-Analysis_Fake"

# subjects = ['Ball', 'Pups', 'Fake', 'Odor'] #As named in Folders

subjects = ['V. Females', 'V. Males Non-Attacking']
Behaviors = ['Attack', 'Sniff', 'Aggressive Behaviors', 'Approach', 'Dead_Pup', 'Grooming', 'Aggressive groom', 'Digging', 
             'Rearing', 'Non-Stimulus Behaviors', 'Nudge', 'Carrying', 'Stimulus Contact', 'Stimulus Interaction', 'qtip', 
             'stick', 'Cotton tip', 'Sniff_InZone', 'Grooming_InZone', 'Sniff_OutZone', 'In nest', 'Qtip Interaction', 
             'Stimulus Interaction_InZone', 'Stimulus Contact_InZone']
use_trace = "Zscore_trace"

def create_excel(subjects, behaviors, compile_file_path):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for behavior in behaviors:
        ws = wb.create_sheet(behavior)
        for col, subject in enumerate(subjects, start=1):
            ws.cell(row=1, column=2 * col - 1, value=f'ID_{subject}')
            ws.cell(row=1, column=2 * col, value=subject)
        # ref_col = len(subjects) * 2 + 1
        # ws.cell(row=1, column=ref_col, value="Reference_ID")
        # ws.cell(row=1, column=ref_col + 1, value="Reference")
    wb.save(compile_file_path)

def write_data_to_workbook(filepath, excel_filename, subject_name):
    df = pd.read_excel(filepath, sheet_name=use_trace, engine='openpyxl')
    wb = openpyxl.load_workbook(excel_filename)
    
    for column in df.columns:
        if column in wb.sheetnames:
            ws = wb[column]
            subject_col = None
            for col_index in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_index).value == subject_name:
                    subject_col = col_index
                    break
            data = df[column]
            data_id = df[f'Animal ID_{column}']
            for row_index, value in enumerate(data, start=2):
                ws.cell(row=row_index, column=subject_col, value=value)
            for row_index, value in enumerate(data_id, start=2):
                ws.cell(row=row_index, column=subject_col - 1, value=value)
    print(f'{subject_name} data written to {filepath}')
    wb.save(excel_filename)

def delete_empty_sheets(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not any(ws.iter_rows(min_row=2, values_only=True)):  # Skip header row
            wb.remove(ws)
    wb.save(filename)

def main():
    rootdir = os.listdir(compile_directory)
    compile_list = []
    for i in rootdir:
        if i in subjects:
            tag = compile_directory.split('\\')[-1].split('_')[-1]
            summary_folder = os.path.join(compile_directory, i, 'Summary', 'during', 'Individual Values')
            if not os.path.exists(summary_folder):
                continue  # Skip if the directory doesn't exist
            summaries = os.listdir(summary_folder)
            for summary in summaries:
                summary_path = os.path.join(summary_folder, summary)
                summary_name = os.path.splitext(summary)[0]
                compile_file_path = os.path.join(compile_directory, f'IndV_ID_{tag}_during_{summary_name}{use_trace}_compiled.xlsx')
                if not os.path.exists(compile_file_path):
                    create_excel(subjects, Behaviors, compile_file_path)
                compile_list.append(compile_file_path)
                write_data_to_workbook(summary_path, compile_file_path, i)
    for compile_file_path in compile_list:
        delete_empty_sheets(compile_file_path)

if __name__ == "__main__":
    main()
